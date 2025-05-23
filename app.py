from flask import Flask, render_template, request, redirect, url_for, session, flash
import pandas as pd
import os
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# File paths
EXCEL_FILE = 'inventory_data.xlsx'
USERS_FILE = 'users.csv'

# Ensure users file exists
if not os.path.exists(USERS_FILE):
    pd.DataFrame(columns=['username', 'password']).to_csv(USERS_FILE, index=False)

# Ensure Excel file with required sheets exists
if not os.path.exists(EXCEL_FILE):
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        pd.DataFrame(columns=['User', 'Date', 'Time', 'Insert Number', 'Quantity']).to_excel(writer, sheet_name='Inward', index=False)
        pd.DataFrame(columns=['User', 'Date', 'Time', 'Insert Number', 'Op Code', 'Tool Number', 'Quantity']).to_excel(writer, sheet_name='Outward', index=False)
        pd.DataFrame(columns=['Insert Number', 'Total Quantity']).to_excel(writer, sheet_name='TotalInventory', index=False)

# Helper functions
def load_users():
    return pd.read_csv(USERS_FILE)

def save_users(df):
    df.to_csv(USERS_FILE, index=False)

def update_total_inventory(insert_number, change):
    wb = load_workbook(EXCEL_FILE)
    sheet = wb['TotalInventory']
    found = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == insert_number:
            row[1].value = (row[1].value or 0) + change
            found = True
            break
    if not found:
        sheet.append([insert_number, change])
    wb.save(EXCEL_FILE)

def read_sheet(sheet_name):
    return pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)

# Routes
@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'login':
            username = request.form['username']
            password = request.form['password']
            users = load_users()
            user = users[users['username'] == username]
            if not user.empty and check_password_hash(user.iloc[0]['password'], password):
                session['username'] = username
                return redirect(url_for('dashboard'))
            else:
                flash('Invalid username or password.')
        elif action == 'register':
            new_username = request.form['new_username']
            new_password = request.form['new_password']
            users = load_users()
            if new_username in users['username'].values:
                flash('Username already exists.')
            else:
                hashed_password = generate_password_hash(new_password)
                new_row = pd.DataFrame([{'username': new_username, 'password': hashed_password}])
                users = pd.concat([users, new_row], ignore_index=True)
                save_users(users)
                flash('User registered successfully.')
        elif action == 'change_password':
            change_username = request.form['change_username']
            current_password = request.form['current_password']
            new_password = request.form['new_password']
            users = load_users()
            user_index = users[users['username'] == change_username].index
            if not user_index.empty and check_password_hash(users.loc[user_index[0], 'password'], current_password):
                users.loc[user_index[0], 'password'] = generate_password_hash(new_password)
                save_users(users)
                flash('Password changed successfully.')
            else:
                flash('Invalid username or current password.')
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('home'))
    total_inventory = read_sheet('TotalInventory')
    return render_template('dashboard.html', username=session['username'], total_inventory=total_inventory.to_dict(orient='records'))

@app.route('/inward', methods=['GET', 'POST'])
def inward():
    if 'username' not in session:
        return redirect(url_for('home'))
    if request.method == 'POST':
        insert_number = request.form['insert_number']
        quantity = int(request.form['quantity'])
        now = datetime.now()
        data = {
            'User': session['username'],
            'Date': now.strftime('%Y-%m-%d'),
            'Time': now.strftime('%H:%M:%S'),
            'Insert Number': insert_number,
            'Quantity': quantity
        }
        df = pd.read_excel(EXCEL_FILE, sheet_name='Inward')
        df = pd.concat([df, pd.DataFrame([data])], ignore_index=True)
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='Inward', index=False)
        update_total_inventory(insert_number, quantity)
    df_inward = read_sheet('Inward')
    total_inventory = read_sheet('TotalInventory')
    return render_template('inward.html', username=session['username'], data=df_inward.to_dict(orient='records'), total_inventory=total_inventory.to_dict(orient='records'))

@app.route('/outward', methods=['GET', 'POST'])
def outward():
    if 'username' not in session:
        return redirect(url_for('home'))
    op_codes = ['OP10','OP20','OP30','OP40','OP50','OP60','OP70','OP80','OP90','OP100']
    if request.method == 'POST':
        insert_number = request.form['insert_number']
        op_code = request.form['op_code']
        tool_number = request.form['tool_number']
        quantity = 1
        now = datetime.now()
        data = {
            'User': session['username'],
            'Date': now.strftime('%Y-%m-%d'),
            'Time': now.strftime('%H:%M:%S'),
            'Insert Number': insert_number,
            'Op Code': op_code,
            'Tool Number': tool_number,
            'Quantity': quantity
        }
        df = pd.read_excel(EXCEL_FILE, sheet_name='Outward')
        df = pd.concat([df, pd.DataFrame([data])], ignore_index=True)
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='Outward', index=False)
        update_total_inventory(insert_number, -quantity)
    df_outward = read_sheet('Outward')
    total_inventory = read_sheet('TotalInventory')
    return render_template('outward.html', username=session['username'], data=df_outward.to_dict(orient='records'), total_inventory=total_inventory.to_dict(orient='records'), op_codes=op_codes)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug=True)
